"""
Transaction Import Module
=========================

Parses CSV/XLSX export files from Coinbase, Kraken, Gemini, and Crypto.com
into normalized Transaction records for the portfolio tracker.

Supported formats:
- Coinbase CSV (2-row header + data)
- Kraken ledger CSV (spend/receive paired by refid)
- Gemini trade XLSX (wide format with per-asset columns)
- Gemini staking XLSX (interest credits)
- Crypto.com CSV (Transaction Kind based)
"""

import csv
import io
import os
import re
import zipfile
from collections import defaultdict
from datetime import datetime
from typing import List, Optional, Tuple

from transaction_history import Transaction

# Try openpyxl for XLSX support
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _strip_currency(value: str) -> float:
    """Strip $ and commas from currency strings, return float."""
    if not value or value == "":
        return 0.0
    s = str(value).replace("$", "").replace(",", "").strip()
    if s == "" or s == "-":
        return 0.0
    return float(s)


def _make_id(exchange: str, timestamp: datetime, asset: str, amount: float, tx_type: str) -> str:
    """Generate a deterministic transaction ID for deduplication."""
    ts_str = timestamp.strftime("%Y%m%d%H%M%S")
    return f"{exchange}_{ts_str}_{asset}_{amount:.8f}_{tx_type}"


# =============================================================================
# Coinbase CSV Parser
# =============================================================================

def import_coinbase_csv(filepath: str) -> List[Transaction]:
    """Parse Coinbase transaction export CSV.

    Format: 2 header rows ('Transactions' line, 'User,name,id' line),
    then column headers, then data rows.
    Columns: ID, Timestamp, Transaction Type, Asset, Quantity Transacted,
             Price Currency, Price at Transaction, Subtotal, Total, Fees, Notes
    """
    transactions = []

    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Find the actual header row (starts with "ID,Timestamp")
    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith("ID,Timestamp"):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"Could not find Coinbase CSV header in {filepath}")

    reader = csv.DictReader(lines[header_idx:])

    for row in reader:
        try:
            raw_type = row.get("Transaction Type", "").strip()
            asset = row.get("Asset", "").strip().upper()
            timestamp_str = row.get("Timestamp", "").strip()
            tx_id = row.get("ID", "").strip()

            if not timestamp_str or not asset:
                continue

            # Parse timestamp: "2025-12-31 22:53:14 UTC"
            timestamp_str = timestamp_str.replace(" UTC", "")
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")

            quantity = float(row.get("Quantity Transacted", "0") or "0")
            price = _strip_currency(row.get("Price at Transaction", "0"))
            subtotal = _strip_currency(row.get("Subtotal", "0"))
            total = _strip_currency(
                row.get("Total (inclusive of fees and/or spread)", "0")
            )
            fees = _strip_currency(row.get("Fees and/or Spread", "0"))

            # Map Coinbase types to normalized types
            type_map = {
                "Buy": "buy",
                "Sell": "sell",
                "Send": "transfer_out",
                "Receive": "transfer_in",
                "Staking Income": "staking_reward",
                "Reward Income": "staking_reward",
                "Incentives Rewards Payout": "reward",
                "Deposit": "deposit",
                "Convert": "convert",
                "Retail Staking Transfer": "stake",
                "Retail Eth2 Deprecation": "transfer_in",
                "Asset Migration": "transfer_in",
                "Raise Offering Deposit": "transfer_out",
                "Raise Offering Distribution": "transfer_in",
                "Raise Offering Pooling": "transfer_out",
            }

            tx_type = type_map.get(raw_type)
            if tx_type is None:
                # Unknown type - skip quietly
                continue

            # Handle Convert: produces two transactions (sell + buy)
            if raw_type == "Convert":
                notes = row.get("Notes", "")
                # Notes like: "Converted 0.5 ETH to 1.2 SOL"
                # The row itself is the "from" side (sell)
                txn_sell = Transaction(
                    id=tx_id or _make_id("Coinbase", timestamp, asset, quantity, "sell"),
                    exchange="Coinbase",
                    timestamp=timestamp,
                    type="sell",
                    asset=asset,
                    amount=abs(quantity),
                    price_usd=price,
                    total_usd=abs(subtotal),
                    fee_usd=abs(fees),
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=abs(subtotal),
                    raw_data=dict(row),
                )
                transactions.append(txn_sell)
                # Try to parse the "to" side from Notes
                match = re.search(
                    r"Converted .+ to ([\d.]+) (\w+)", notes
                )
                if match:
                    to_amount = float(match.group(1))
                    to_asset = match.group(2).upper()
                    txn_buy = Transaction(
                        id=f"{tx_id}_convert_buy" if tx_id else _make_id(
                            "Coinbase", timestamp, to_asset, to_amount, "buy"
                        ),
                        exchange="Coinbase",
                        timestamp=timestamp,
                        type="buy",
                        asset=to_asset,
                        amount=to_amount,
                        price_usd=abs(subtotal) / to_amount if to_amount else 0,
                        total_usd=abs(subtotal),
                        fee_usd=0.0,
                        fee_asset="USD",
                        related_asset=asset,
                        related_amount=abs(quantity),
                        raw_data=dict(row),
                    )
                    transactions.append(txn_buy)
                continue

            # Handle negative quantities for staking transfers
            if quantity < 0 and tx_type == "stake":
                # Negative = unstaking side of the pair, skip (paired with positive)
                continue

            txn = Transaction(
                id=tx_id or _make_id("Coinbase", timestamp, asset, abs(quantity), tx_type),
                exchange="Coinbase",
                timestamp=timestamp,
                type=tx_type,
                asset=asset,
                amount=abs(quantity),
                price_usd=price,
                total_usd=abs(total) if total else abs(subtotal),
                fee_usd=abs(fees),
                fee_asset="USD",
                related_asset="USD",
                related_amount=abs(total) if total else abs(subtotal),
                raw_data=dict(row),
            )
            transactions.append(txn)

        except Exception as e:
            # Skip malformed rows
            continue

    return transactions


# =============================================================================
# Kraken Ledger CSV Parser
# =============================================================================

KRAKEN_ASSET_MAP = {
    "USD.HOLD": "USD",
    "USD.M": "USD",
    "ZUSD": "USD",
    "XXBT": "BTC",
    "XETH": "ETH",
    "XXRP": "XRP",
    "XLTC": "LTC",
    "XXLM": "XLM",
    "XDOGE": "DOGE",
    "XXDG": "DOGE",
}


def _normalize_kraken_asset(raw: str) -> str:
    """Normalize Kraken asset symbols to standard tickers."""
    raw = raw.strip()
    if raw in KRAKEN_ASSET_MAP:
        return KRAKEN_ASSET_MAP[raw]
    # Strip leading X/Z for 4-char Kraken symbols
    if len(raw) == 4 and raw[0] in ("X", "Z"):
        return raw[1:]
    return raw.upper()


def import_kraken_ledger(filepath: str) -> List[Transaction]:
    """Parse Kraken ledger CSV export.

    Columns: txid, refid, time, type, subtype, aclass, subclass, asset,
             wallet, amount, fee, balance, amountusd
    Trades are paired as spend+receive sharing the same refid.
    """
    transactions = []

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Group by refid to pair spend/receive entries
    by_refid = defaultdict(list)
    standalone = []

    for row in rows:
        tx_type = row.get("type", "").strip()
        refid = row.get("refid", "").strip()

        if tx_type in ("spend", "receive") and refid:
            by_refid[refid].append(row)
        else:
            standalone.append(row)

    # Process paired trades (spend + receive = one trade)
    for refid, entries in by_refid.items():
        try:
            spend_row = None
            receive_row = None
            for e in entries:
                if e["type"].strip() == "spend":
                    spend_row = e
                elif e["type"].strip() == "receive":
                    receive_row = e

            if not spend_row or not receive_row:
                # Unpaired - treat individually
                for e in entries:
                    _add_kraken_standalone(e, transactions)
                continue

            spend_asset = _normalize_kraken_asset(spend_row["asset"])
            receive_asset = _normalize_kraken_asset(receive_row["asset"])
            spend_amount = abs(float(spend_row["amount"]))
            receive_amount = abs(float(receive_row["amount"]))
            spend_fee = abs(float(spend_row["fee"]))
            receive_fee = abs(float(receive_row["fee"]))
            timestamp_str = spend_row["time"].strip()
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
            spend_usd = abs(float(spend_row.get("amountusd", "0") or "0"))
            receive_usd = abs(float(receive_row.get("amountusd", "0") or "0"))

            # Determine which side is the "base" (crypto) vs "quote" (USD/stablecoin)
            if spend_asset in ("USD", "USDT", "USDC", "DAI"):
                # Spending USD to buy crypto
                txn = Transaction(
                    id=refid,
                    exchange="Kraken",
                    timestamp=timestamp,
                    type="buy",
                    asset=receive_asset,
                    amount=receive_amount,
                    price_usd=spend_amount / receive_amount if receive_amount else 0,
                    total_usd=spend_amount,
                    fee_usd=spend_fee + receive_fee,
                    fee_asset=spend_asset,
                    related_asset=spend_asset,
                    related_amount=spend_amount,
                    raw_data={"spend": dict(spend_row), "receive": dict(receive_row)},
                )
            elif receive_asset in ("USD", "USDT", "USDC", "DAI"):
                # Receiving USD from selling crypto
                txn = Transaction(
                    id=refid,
                    exchange="Kraken",
                    timestamp=timestamp,
                    type="sell",
                    asset=spend_asset,
                    amount=spend_amount,
                    price_usd=receive_amount / spend_amount if spend_amount else 0,
                    total_usd=receive_amount,
                    fee_usd=spend_fee + receive_fee,
                    fee_asset=receive_asset,
                    related_asset=receive_asset,
                    related_amount=receive_amount,
                    raw_data={"spend": dict(spend_row), "receive": dict(receive_row)},
                )
            else:
                # Crypto-to-crypto trade: record as sell + buy
                txn_sell = Transaction(
                    id=f"{refid}_sell",
                    exchange="Kraken",
                    timestamp=timestamp,
                    type="sell",
                    asset=spend_asset,
                    amount=spend_amount,
                    price_usd=spend_usd / spend_amount if spend_amount else 0,
                    total_usd=spend_usd,
                    fee_usd=spend_fee,
                    fee_asset=spend_asset,
                    related_asset=receive_asset,
                    related_amount=receive_amount,
                    raw_data=dict(spend_row),
                )
                txn_buy = Transaction(
                    id=f"{refid}_buy",
                    exchange="Kraken",
                    timestamp=timestamp,
                    type="buy",
                    asset=receive_asset,
                    amount=receive_amount,
                    price_usd=receive_usd / receive_amount if receive_amount else 0,
                    total_usd=receive_usd,
                    fee_usd=receive_fee,
                    fee_asset=receive_asset,
                    related_asset=spend_asset,
                    related_amount=spend_amount,
                    raw_data=dict(receive_row),
                )
                transactions.append(txn_sell)
                transactions.append(txn_buy)
                continue

            transactions.append(txn)

        except Exception:
            continue

    # Process standalone entries (deposits, withdrawals, earn, transfers)
    for row in standalone:
        _add_kraken_standalone(row, transactions)

    return transactions


def _add_kraken_standalone(row: dict, transactions: List[Transaction]):
    """Process a standalone (non-paired) Kraken ledger entry."""
    try:
        tx_type = row.get("type", "").strip()
        asset = _normalize_kraken_asset(row.get("asset", ""))
        amount = float(row.get("amount", "0"))
        fee = abs(float(row.get("fee", "0")))
        timestamp_str = row.get("time", "").strip()
        amount_usd = abs(float(row.get("amountusd", "0") or "0"))
        txid = row.get("txid", "").strip()
        refid = row.get("refid", "").strip()

        if not timestamp_str or not asset:
            return

        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")

        type_map = {
            "deposit": "deposit",
            "withdrawal": "transfer_out",
            "earn": "staking_reward",
            "transfer": "transfer_in",
        }

        normalized_type = type_map.get(tx_type)
        if normalized_type is None:
            return

        # For withdrawals, amount is negative
        abs_amount = abs(amount)

        txn = Transaction(
            id=txid or refid or _make_id("Kraken", timestamp, asset, abs_amount, normalized_type),
            exchange="Kraken",
            timestamp=timestamp,
            type=normalized_type,
            asset=asset,
            amount=abs_amount,
            price_usd=amount_usd / abs_amount if abs_amount else 0,
            total_usd=amount_usd,
            fee_usd=fee,
            fee_asset=asset,
            related_asset="USD",
            related_amount=amount_usd,
            raw_data=dict(row),
        )
        transactions.append(txn)

    except Exception:
        pass


# =============================================================================
# Gemini XLSX Parser (Trade History)
# =============================================================================

def import_gemini_xlsx(filepath: str) -> List[Transaction]:
    """Parse Gemini trade history XLSX (wide format).

    Wide format: per-asset Amount/Fee/Balance columns after the fixed columns.
    Fixed columns: Date, Time (UTC), Type, Symbol, Specification, Liquidity Indicator,
                   Trading Fee Rate (bps), USD Amount USD, Fee (USD) USD, USD Balance USD
    Then groups of 3: {ASSET} Amount {ASSET}, Fee ({ASSET}) {ASSET}, {ASSET} Balance {ASSET}
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for XLSX parsing: pip install openpyxl")

    transactions = []
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Read header row
    header = [cell.value for cell in ws[1]]

    # Identify asset columns (groups of 3: Amount, Fee, Balance)
    asset_columns = {}  # asset -> (amount_col_idx, fee_col_idx, balance_col_idx)
    for i, h in enumerate(header):
        if h and "Amount" in str(h) and i >= 10:  # Skip USD columns at 7-9
            # Pattern: "{ASSET} Amount {ASSET}" or "Amount {ASSET}"
            parts = str(h).split()
            asset_name = parts[-1] if parts else ""
            if asset_name and asset_name != "USD":
                asset_columns[asset_name] = (i, i + 1, i + 2)

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            date_val = row[0]
            time_val = row[1]
            tx_type = row[2]
            symbol = row[3]

            if not date_val or not tx_type:
                continue

            # Build timestamp from Date + Time columns
            if isinstance(date_val, datetime):
                timestamp = date_val
            else:
                continue

            usd_amount = abs(float(row[7] or 0))
            usd_fee = abs(float(row[8] or 0))

            type_map = {
                "Buy": "buy",
                "Sell": "sell",
                "Credit": "deposit",
                "Debit": "transfer_out",
            }
            normalized_type = type_map.get(tx_type)
            if normalized_type is None:
                continue

            # For Buy/Sell, find which asset column has a non-zero amount
            if normalized_type in ("buy", "sell") and symbol:
                # Parse trading pair from Symbol (e.g., "SHIBUSD", "BTCUSD")
                # The crypto asset is everything except the last 3 chars (USD)
                pair_asset = None
                crypto_amount = 0.0
                asset_fee = 0.0

                if symbol.endswith("USD") and len(symbol) > 3:
                    pair_asset = symbol[:-3].upper()
                elif symbol.endswith("GUSD") and len(symbol) > 4:
                    pair_asset = symbol[:-4].upper()

                # Look up in asset columns
                if pair_asset and pair_asset in asset_columns:
                    amt_idx = asset_columns[pair_asset][0]
                    fee_idx = asset_columns[pair_asset][1]
                    crypto_amount = abs(float(row[amt_idx] or 0))
                    asset_fee = abs(float(row[fee_idx] or 0))

                if crypto_amount > 0 and pair_asset:
                    price = usd_amount / crypto_amount if crypto_amount else 0
                    txn = Transaction(
                        id=str(row[22] or "") if len(row) > 22 else _make_id(
                            "Gemini", timestamp, pair_asset, crypto_amount, normalized_type
                        ),
                        exchange="Gemini",
                        timestamp=timestamp,
                        type=normalized_type,
                        asset=pair_asset,
                        amount=crypto_amount,
                        price_usd=price,
                        total_usd=usd_amount,
                        fee_usd=usd_fee,
                        fee_asset="USD",
                        related_asset="USD",
                        related_amount=usd_amount,
                        raw_data={"symbol": symbol, "type": tx_type},
                    )
                    transactions.append(txn)

            elif normalized_type in ("deposit", "transfer_out"):
                # Credit/Debit - find the asset with a non-zero amount
                # Could be USD or any crypto
                if usd_amount > 0 and (not symbol or symbol == "USD"):
                    txn = Transaction(
                        id=_make_id("Gemini", timestamp, "USD", usd_amount, normalized_type),
                        exchange="Gemini",
                        timestamp=timestamp,
                        type=normalized_type,
                        asset="USD",
                        amount=usd_amount,
                        price_usd=1.0,
                        total_usd=usd_amount,
                        fee_usd=usd_fee,
                        fee_asset="USD",
                        related_asset="USD",
                        related_amount=usd_amount,
                        raw_data={"type": tx_type, "symbol": symbol},
                    )
                    transactions.append(txn)
                else:
                    # Find non-zero crypto column
                    for asset_name, (amt_idx, fee_idx, _) in asset_columns.items():
                        amt = row[amt_idx]
                        if amt and abs(float(amt)) > 0:
                            abs_amt = abs(float(amt))
                            txn = Transaction(
                                id=_make_id(
                                    "Gemini", timestamp, asset_name, abs_amt, normalized_type
                                ),
                                exchange="Gemini",
                                timestamp=timestamp,
                                type=normalized_type,
                                asset=asset_name,
                                amount=abs_amt,
                                price_usd=usd_amount / abs_amt if abs_amt and usd_amount else 0,
                                total_usd=usd_amount,
                                fee_usd=usd_fee,
                                fee_asset="USD",
                                related_asset="USD",
                                related_amount=usd_amount,
                                raw_data={"type": tx_type, "symbol": symbol or asset_name},
                            )
                            transactions.append(txn)
                            break

        except Exception:
            continue

    wb.close()
    return transactions


# =============================================================================
# Gemini Staking XLSX Parser
# =============================================================================

def import_gemini_staking_xlsx(filepath: str) -> List[Transaction]:
    """Parse Gemini staking transaction history XLSX.

    Columns: Date, Time (UTC), Type, Symbol, Borrower, APY, Rate (bps), [blank],
             Amount {ASSET}, Price USD, Amount USD, Monthly Summary {ASSET}, Balance {ASSET}
    Types: Deposit, Interest Credit, Withdrawal, Redeem,
           Monthly Interest Summary (skip), Administrative Debit (skip)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for XLSX parsing: pip install openpyxl")

    transactions = []
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    # Detect asset from header: "Amount MATIC" -> "MATIC"
    asset_symbol = None
    for h in header:
        if h and str(h).startswith("Amount ") and not str(h).startswith("Amount USD"):
            asset_symbol = str(h).split()[-1].upper()
            break

    if not asset_symbol:
        asset_symbol = "UNKNOWN"

    # Column indices (0-based)
    # 8: Amount {ASSET}, 9: Price USD, 10: Amount USD
    amt_idx = 8
    price_idx = 9
    usd_idx = 10

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            date_val = row[0]
            tx_type = row[2]

            if not date_val or not tx_type:
                continue

            if isinstance(date_val, datetime):
                timestamp = date_val
            else:
                continue

            # Skip summary and near-zero administrative rows
            if tx_type in ("Monthly Interest Summary",):
                continue

            amount = float(row[amt_idx] or 0)
            price_usd = float(row[price_idx] or 0)
            amount_usd = float(row[usd_idx] or 0)

            if abs(amount) < 1e-15:
                continue

            type_map = {
                "Deposit": "stake",
                "Interest Credit": "staking_reward",
                "Withdrawal": "transfer_out",
                "Redeem": "transfer_in",
                "Administrative Debit": "fee",
            }

            normalized_type = type_map.get(tx_type)
            if normalized_type is None:
                continue

            txn = Transaction(
                id=_make_id("Gemini", timestamp, asset_symbol, abs(amount), normalized_type),
                exchange="Gemini",
                timestamp=timestamp,
                type=normalized_type,
                asset=asset_symbol,
                amount=abs(amount),
                price_usd=price_usd,
                total_usd=abs(amount_usd),
                fee_usd=0.0,
                fee_asset=asset_symbol,
                related_asset="USD",
                related_amount=abs(amount_usd),
                raw_data={"type": tx_type, "symbol": asset_symbol},
            )
            transactions.append(txn)

        except Exception:
            continue

    wb.close()
    return transactions


# =============================================================================
# Crypto.com CSV Parser
# =============================================================================

def import_crypto_com_csv(filepath: str) -> List[Transaction]:
    """Parse Crypto.com transaction export CSV.

    Columns: Timestamp (UTC), Transaction Description, Currency, Amount,
             To Currency, To Amount, Native Currency, Native Amount,
             Native Amount (in USD), Transaction Kind, Transaction Hash
    """
    transactions = []

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Group swap pairs (credited + debited share same timestamp)
    swap_debits = {}  # timestamp -> row

    for row in rows:
        try:
            kind = row.get("Transaction Kind", "").strip()
            timestamp_str = row.get("Timestamp (UTC)", "").strip()
            currency = row.get("Currency", "").strip().upper()
            amount = float(row.get("Amount", "0") or "0")
            to_currency = row.get("To Currency", "").strip().upper()
            to_amount = float(row.get("To Amount", "0") or "0")
            native_usd = abs(float(row.get("Native Amount (in USD)", "0") or "0"))
            tx_hash = row.get("Transaction Hash", "").strip()

            if not timestamp_str or not currency:
                continue

            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")

            # Categorize transaction kinds
            if kind in ("crypto_purchase", "recurring_buy_order", "viban_purchase"):
                txn = Transaction(
                    id=tx_hash or _make_id("Crypto.com", timestamp, currency, abs(amount), "buy"),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="buy",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind == "crypto_exchange":
                # Exchange: sell currency, buy to_currency
                if to_currency and to_amount:
                    txn_sell = Transaction(
                        id=(tx_hash or _make_id(
                            "Crypto.com", timestamp, currency, abs(amount), "sell"
                        )) + "_sell",
                        exchange="Crypto.com",
                        timestamp=timestamp,
                        type="sell",
                        asset=currency,
                        amount=abs(amount),
                        price_usd=native_usd / abs(amount) if amount else 0,
                        total_usd=native_usd,
                        fee_usd=0.0,
                        fee_asset="USD",
                        related_asset=to_currency,
                        related_amount=abs(to_amount),
                        raw_data=dict(row),
                    )
                    txn_buy = Transaction(
                        id=(tx_hash or _make_id(
                            "Crypto.com", timestamp, to_currency, abs(to_amount), "buy"
                        )) + "_buy",
                        exchange="Crypto.com",
                        timestamp=timestamp,
                        type="buy",
                        asset=to_currency,
                        amount=abs(to_amount),
                        price_usd=native_usd / abs(to_amount) if to_amount else 0,
                        total_usd=native_usd,
                        fee_usd=0.0,
                        fee_asset="USD",
                        related_asset=currency,
                        related_amount=abs(amount),
                        raw_data=dict(row),
                    )
                    transactions.append(txn_sell)
                    transactions.append(txn_buy)

            elif kind == "crypto_withdrawal":
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "transfer_out"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="transfer_out",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset=currency,
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind == "crypto_wallet_swap_credited":
                # Buy side of a swap
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "buy"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="buy",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind == "crypto_wallet_swap_debited":
                # Sell side of a swap
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "sell"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="sell",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind.startswith("finance.dpos.non_compound_interest") or \
                 kind.startswith("finance.defi_staking.non_compound_interest") or \
                 kind.startswith("finance.defi_lending.compound_interest"):
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "staking_reward"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="staking_reward",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset=currency,
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind.startswith("finance.dpos.staking") or \
                 kind.startswith("finance.defi_staking.staking") or \
                 kind.startswith("finance.defi_lending.staking"):
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "stake"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="stake",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset=currency,
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind in ("viban_deposit_precredit", "viban_deposit_precredit_repayment"):
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "deposit"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="deposit",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=1.0 if currency in ("USD", "USDT", "USDC") else (
                        native_usd / abs(amount) if amount else 0
                    ),
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind.startswith("reward.loyalty_program"):
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), "reward"
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type="reward",
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset=currency,
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            elif kind.startswith("trading.knock_outs"):
                # Knock-out trading orders - treat as buy/sell based on amount sign
                tx_type_ko = "buy" if amount > 0 else "sell"
                txn = Transaction(
                    id=tx_hash or _make_id(
                        "Crypto.com", timestamp, currency, abs(amount), tx_type_ko
                    ),
                    exchange="Crypto.com",
                    timestamp=timestamp,
                    type=tx_type_ko,
                    asset=currency,
                    amount=abs(amount),
                    price_usd=native_usd / abs(amount) if amount else 0,
                    total_usd=native_usd,
                    fee_usd=0.0,
                    fee_asset="USD",
                    related_asset="USD",
                    related_amount=native_usd,
                    raw_data=dict(row),
                )
                transactions.append(txn)

            # else: skip unknown kinds

        except Exception:
            continue

    return transactions


# =============================================================================
# Auto-detection and unified import
# =============================================================================

def detect_format(filepath: str) -> Optional[str]:
    """Auto-detect exchange format from file contents or name.

    Returns: 'coinbase', 'kraken', 'gemini', 'gemini_staking', 'crypto_com', or None
    """
    basename = os.path.basename(filepath).lower()

    # Name-based detection
    if "coinbase" in basename:
        return "coinbase"
    if "kraken" in basename or "ledger" in basename:
        return "kraken"
    if "staking" in basename and basename.endswith(".xlsx"):
        return "gemini_staking"
    if "crypto_com" in basename:
        return "crypto_com"

    # Content-based detection
    if filepath.endswith(".xlsx"):
        if OPENPYXL_AVAILABLE:
            try:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                title = ws.title or ""
                header = [str(cell.value or "") for cell in ws[1]]
                wb.close()
                if "Staking" in title or any("APY" in h for h in header):
                    return "gemini_staking"
                if "Account History" in title or any("Trading Fee Rate" in h for h in header):
                    return "gemini"
            except Exception:
                pass
        return "gemini"  # Default for .xlsx

    if filepath.endswith(".csv"):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                first_lines = [f.readline() for _ in range(5)]
            combined = "".join(first_lines)

            if "Transactions" in first_lines[0] and "User," in first_lines[1]:
                return "coinbase"
            if "txid" in combined and "refid" in combined:
                return "kraken"
            if "Transaction Kind" in combined or "Native Amount (in USD)" in combined:
                return "crypto_com"
        except Exception:
            pass

    return None


def import_file(filepath: str, exchange_override: Optional[str] = None) -> List[Transaction]:
    """Import transactions from a single file.

    Args:
        filepath: Path to CSV or XLSX file
        exchange_override: Force a specific parser ('coinbase', 'kraken',
                          'gemini', 'gemini_staking', 'crypto_com')

    Returns:
        List of Transaction objects
    """
    fmt = exchange_override or detect_format(filepath)

    if fmt is None:
        raise ValueError(f"Could not detect format for {filepath}")

    parser_map = {
        "coinbase": import_coinbase_csv,
        "kraken": import_kraken_ledger,
        "gemini": import_gemini_xlsx,
        "gemini_staking": import_gemini_staking_xlsx,
        "crypto_com": import_crypto_com_csv,
    }

    parser = parser_map.get(fmt)
    if parser is None:
        raise ValueError(f"Unknown format: {fmt}")

    return parser(filepath)


def deduplicate(transactions: List[Transaction]) -> List[Transaction]:
    """Remove duplicate transactions based on (exchange, timestamp, asset, amount, type)."""
    seen = set()
    unique = []

    for txn in transactions:
        key = (
            txn.exchange,
            txn.timestamp.strftime("%Y%m%d%H%M%S"),
            txn.asset,
            f"{txn.amount:.8f}",
            txn.type,
        )
        if key not in seen:
            seen.add(key)
            unique.append(txn)

    return unique


def import_all(file_paths: List[str], exchange_override: Optional[str] = None) -> Tuple[List[Transaction], dict]:
    """Import and merge transactions from multiple files.

    Args:
        file_paths: List of CSV/XLSX file paths to import
        exchange_override: Force parser for all files

    Returns:
        Tuple of (deduplicated transactions sorted by timestamp, summary dict)
    """
    all_transactions = []
    summary = {
        "files_processed": 0,
        "files_failed": [],
        "per_file": {},
        "per_exchange": defaultdict(int),
        "per_type": defaultdict(int),
        "total_raw": 0,
        "total_deduped": 0,
        "date_range": {"earliest": None, "latest": None},
    }

    for filepath in file_paths:
        try:
            txns = import_file(filepath, exchange_override)
            basename = os.path.basename(filepath)
            summary["files_processed"] += 1
            summary["per_file"][basename] = len(txns)
            summary["total_raw"] += len(txns)

            for txn in txns:
                summary["per_exchange"][txn.exchange] += 1
                summary["per_type"][txn.type] += 1

            all_transactions.extend(txns)

        except Exception as e:
            summary["files_failed"].append({"file": filepath, "error": str(e)})

    # Deduplicate and sort
    all_transactions = deduplicate(all_transactions)
    all_transactions.sort(key=lambda x: x.timestamp)

    summary["total_deduped"] = len(all_transactions)

    if all_transactions:
        summary["date_range"]["earliest"] = all_transactions[0].timestamp.isoformat()
        summary["date_range"]["latest"] = all_transactions[-1].timestamp.isoformat()

    # Convert defaultdicts for JSON serialization
    summary["per_exchange"] = dict(summary["per_exchange"])
    summary["per_type"] = dict(summary["per_type"])

    return all_transactions, summary
