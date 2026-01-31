"""
Portfolio Export
Export combined crypto portfolio to CSV, Excel, or JSON formats.

Usage:
    python export_portfolio.py                    # Export to CSV
    python export_portfolio.py --format excel     # Export to Excel
    python export_portfolio.py --format json      # Export to JSON
    python export_portfolio.py --include-staking  # Include staking details
"""

import os
import csv
import json
import argparse
from datetime import datetime
from typing import Dict, List, Optional
from dotenv import load_dotenv

# Import exchange clients
from exchanges import CoinbaseClient, KrakenClient, CryptoComClient, GeminiClient


class PortfolioExporter:
    """Export portfolio data from multiple exchanges."""
    
    EXCHANGE_NAMES = {
        "coinbase": "Coinbase",
        "kraken": "Kraken",
        "cryptocom": "Crypto.com",
        "gemini": "Gemini"
    }
    
    def __init__(self):
        load_dotenv()
        self.clients: Dict[str, object] = {}
        self.portfolio_data: List[dict] = []
        self.staking_data: List[dict] = []
        self.summary: dict = {}
        
    def connect_all(self) -> int:
        """Connect to all configured exchanges."""
        connected = 0
        
        # Coinbase
        key = os.getenv("COINBASE_API_KEY")
        secret = os.getenv("COINBASE_API_SECRET")
        if key and secret:
            self.clients["coinbase"] = CoinbaseClient(key, secret)
            connected += 1
        
        # Kraken
        key = os.getenv("KRAKEN_API_KEY")
        secret = os.getenv("KRAKEN_API_SECRET")
        if key and secret:
            client = KrakenClient(key, secret)
            if client.get_server_time():
                self.clients["kraken"] = client
                connected += 1
        
        # Crypto.com
        key = os.getenv("CRYPTOCOM_API_KEY")
        secret = os.getenv("CRYPTOCOM_API_SECRET")
        if key and secret:
            self.clients["cryptocom"] = CryptoComClient(key, secret)
            connected += 1
        
        # Gemini
        key = os.getenv("GEMINI_API_KEY")
        secret = os.getenv("GEMINI_API_SECRET")
        if key and secret:
            self.clients["gemini"] = GeminiClient(key, secret)
            connected += 1
        
        return connected
    
    def fetch_all_data(self, include_staking: bool = True):
        """Fetch portfolio data from all connected exchanges."""
        self.portfolio_data = []
        self.staking_data = []
        
        timestamp = datetime.now().isoformat()
        total_value = 0
        total_staked = 0
        
        for exchange, client in self.clients.items():
            try:
                accounts = client.get_accounts()
                
                for account in accounts:
                    currency = account.get("currency", "UNKNOWN")
                    balance = float(account.get("available_balance", {}).get("value", 0))
                    hold = float(account.get("hold", {}).get("value", 0))
                    
                    if balance <= 0 and hold <= 0:
                        continue
                    
                    # Determine if staked
                    is_staked = account.get("is_staked", False)
                    if "(Staked)" in currency:
                        is_staked = True
                        currency = currency.replace(" (Staked)", "")
                    
                    # Get current price
                    price = client.get_spot_price(currency, "USD")
                    
                    # Handle stablecoins and fiat
                    if currency in ["USD", "USDC", "USDT", "ZUSD"]:
                        price = 1.0
                        currency = currency.replace("Z", "")  # Normalize Kraken's ZUSD
                    
                    usd_value = (balance + hold) * price if price else 0
                    
                    # Get price changes
                    changes = None
                    if price and hasattr(client, 'get_price_changes'):
                        changes = client.get_price_changes(currency, "USD")
                    
                    row = {
                        "timestamp": timestamp,
                        "exchange": self.EXCHANGE_NAMES.get(exchange, exchange),
                        "asset": currency,
                        "balance": balance,
                        "on_hold": hold,
                        "total_balance": balance + hold,
                        "price_usd": price,
                        "value_usd": usd_value,
                        "is_staked": is_staked,
                        "change_24h": changes.get("change_24h") if changes else None,
                        "change_7d": changes.get("change_7d") if changes else None,
                        "change_30d": changes.get("change_30d") if changes else None,
                    }
                    
                    # Additional staking info if available
                    if is_staked:
                        row["apy"] = account.get("apy")
                        row["accrued_interest"] = account.get("accrued_interest")
                        self.staking_data.append(row)
                        total_staked += usd_value
                    
                    self.portfolio_data.append(row)
                    total_value += usd_value
                    
            except Exception as e:
                print(f"Error fetching from {exchange}: {e}")
        
        # Create summary
        self.summary = {
            "timestamp": timestamp,
            "total_value_usd": total_value,
            "total_staked_usd": total_staked,
            "total_liquid_usd": total_value - total_staked,
            "num_assets": len(set(r["asset"] for r in self.portfolio_data)),
            "num_exchanges": len(self.clients),
            "exchanges": list(self.EXCHANGE_NAMES.get(ex, ex) for ex in self.clients.keys())
        }
        
        # Calculate per-exchange totals
        exchange_totals = {}
        for row in self.portfolio_data:
            ex = row["exchange"]
            exchange_totals[ex] = exchange_totals.get(ex, 0) + row["value_usd"]
        self.summary["exchange_totals"] = exchange_totals
        
        # Calculate per-asset totals
        asset_totals = {}
        for row in self.portfolio_data:
            asset = row["asset"]
            if asset not in asset_totals:
                asset_totals[asset] = {
                    "total_balance": 0,
                    "total_value_usd": 0,
                    "exchanges": []
                }
            asset_totals[asset]["total_balance"] += row["total_balance"]
            asset_totals[asset]["total_value_usd"] += row["value_usd"]
            asset_totals[asset]["exchanges"].append(row["exchange"])
        self.summary["asset_totals"] = asset_totals
    
    def export_csv(self, filename: str = None) -> str:
        """Export portfolio to CSV file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"portfolio_export_{timestamp}.csv"
        
        if not self.portfolio_data:
            self.fetch_all_data()
        
        fieldnames = [
            "timestamp", "exchange", "asset", "balance", "on_hold", 
            "total_balance", "price_usd", "value_usd", "is_staked",
            "change_24h", "change_7d", "change_30d", "apy", "accrued_interest"
        ]
        
        with open(filename, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(self.portfolio_data)
        
        print(f"✅ Exported to {filename}")
        return filename
    
    def export_excel(self, filename: str = None) -> str:
        """Export portfolio to Excel file with multiple sheets."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            print("❌ openpyxl not installed. Run: pip install openpyxl")
            return None
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"portfolio_export_{timestamp}.xlsx"
        
        if not self.portfolio_data:
            self.fetch_all_data()
        
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        money_format = '$#,##0.00'
        pct_format = '0.00%'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Portfolio Summary", ""],
            ["Generated", self.summary["timestamp"]],
            ["", ""],
            ["Total Value", self.summary["total_value_usd"]],
            ["Total Staked", self.summary["total_staked_usd"]],
            ["Total Liquid", self.summary["total_liquid_usd"]],
            ["", ""],
            ["Number of Assets", self.summary["num_assets"]],
            ["Number of Exchanges", self.summary["num_exchanges"]],
            ["", ""],
            ["Value by Exchange", ""],
        ]
        
        for ex, val in self.summary.get("exchange_totals", {}).items():
            summary_data.append([ex, val])
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                if col_idx == 2 and isinstance(value, (int, float)) and value > 100:
                    cell.number_format = money_format
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Sheet 2: All Holdings
        ws_holdings = wb.create_sheet("Holdings")
        
        headers = ["Exchange", "Asset", "Balance", "On Hold", "Price (USD)", 
                   "Value (USD)", "Staked", "24h %", "7d %", "30d %"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_holdings.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row_data in enumerate(self.portfolio_data, 2):
            ws_holdings.cell(row=row_idx, column=1, value=row_data["exchange"])
            ws_holdings.cell(row=row_idx, column=2, value=row_data["asset"])
            ws_holdings.cell(row=row_idx, column=3, value=row_data["balance"])
            ws_holdings.cell(row=row_idx, column=4, value=row_data["on_hold"])
            
            price_cell = ws_holdings.cell(row=row_idx, column=5, value=row_data["price_usd"])
            price_cell.number_format = money_format
            
            value_cell = ws_holdings.cell(row=row_idx, column=6, value=row_data["value_usd"])
            value_cell.number_format = money_format
            
            ws_holdings.cell(row=row_idx, column=7, value="Yes" if row_data["is_staked"] else "No")
            
            for col, key in [(8, "change_24h"), (9, "change_7d"), (10, "change_30d")]:
                value = row_data.get(key)
                if value is not None:
                    cell = ws_holdings.cell(row=row_idx, column=col, value=value/100)
                    cell.number_format = pct_format
                    if value < 0:
                        cell.font = Font(color="FF0000")
                    else:
                        cell.font = Font(color="00AA00")
        
        # Auto-adjust column widths
        for col in ws_holdings.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_holdings.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
        
        # Sheet 3: By Asset (aggregated)
        ws_assets = wb.create_sheet("By Asset")
        
        asset_headers = ["Asset", "Total Balance", "Total Value (USD)", "% of Portfolio", "Exchanges"]
        for col, header in enumerate(asset_headers, 1):
            cell = ws_assets.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        total_value = self.summary["total_value_usd"]
        sorted_assets = sorted(
            self.summary.get("asset_totals", {}).items(),
            key=lambda x: x[1]["total_value_usd"],
            reverse=True
        )
        
        for row_idx, (asset, data) in enumerate(sorted_assets, 2):
            ws_assets.cell(row=row_idx, column=1, value=asset)
            ws_assets.cell(row=row_idx, column=2, value=data["total_balance"])
            
            value_cell = ws_assets.cell(row=row_idx, column=3, value=data["total_value_usd"])
            value_cell.number_format = money_format
            
            pct = data["total_value_usd"] / total_value if total_value > 0 else 0
            pct_cell = ws_assets.cell(row=row_idx, column=4, value=pct)
            pct_cell.number_format = pct_format
            
            ws_assets.cell(row=row_idx, column=5, value=", ".join(set(data["exchanges"])))
        
        for col in ws_assets.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_assets.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)
        
        # Sheet 4: Staking (if any)
        if self.staking_data:
            ws_staking = wb.create_sheet("Staking")
            
            staking_headers = ["Exchange", "Asset", "Balance", "Value (USD)", "APY", "Accrued Interest"]
            for col, header in enumerate(staking_headers, 1):
                cell = ws_staking.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row_idx, row_data in enumerate(self.staking_data, 2):
                ws_staking.cell(row=row_idx, column=1, value=row_data["exchange"])
                ws_staking.cell(row=row_idx, column=2, value=row_data["asset"])
                ws_staking.cell(row=row_idx, column=3, value=row_data["total_balance"])
                
                value_cell = ws_staking.cell(row=row_idx, column=4, value=row_data["value_usd"])
                value_cell.number_format = money_format
                
                apy = row_data.get("apy")
                if apy:
                    apy_cell = ws_staking.cell(row=row_idx, column=5, value=apy/100 if apy > 1 else apy)
                    apy_cell.number_format = pct_format
                
                accrued = row_data.get("accrued_interest")
                if accrued:
                    ws_staking.cell(row=row_idx, column=6, value=accrued)
        
        wb.save(filename)
        print(f"✅ Exported to {filename}")
        return filename
    
    def export_json(self, filename: str = None, pretty: bool = True) -> str:
        """Export portfolio to JSON file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"portfolio_export_{timestamp}.json"
        
        if not self.portfolio_data:
            self.fetch_all_data()
        
        export_data = {
            "summary": self.summary,
            "holdings": self.portfolio_data,
            "staking": self.staking_data
        }
        
        with open(filename, 'w') as f:
            if pretty:
                json.dump(export_data, f, indent=2, default=str)
            else:
                json.dump(export_data, f, default=str)
        
        print(f"✅ Exported to {filename}")
        return filename
    
    def print_summary(self):
        """Print portfolio summary to console."""
        if not self.summary:
            self.fetch_all_data()
        
        print(f"\n{'='*60}")
        print(f"📊 PORTFOLIO SUMMARY")
        print(f"{'='*60}")
        print(f"Generated: {self.summary['timestamp']}")
        print(f"\n💰 Total Value: ${self.summary['total_value_usd']:,.2f}")
        print(f"🔒 Staked:      ${self.summary['total_staked_usd']:,.2f}")
        print(f"💧 Liquid:      ${self.summary['total_liquid_usd']:,.2f}")
        print(f"\n📈 Assets:    {self.summary['num_assets']}")
        print(f"🏦 Exchanges: {self.summary['num_exchanges']}")
        
        print(f"\n📊 By Exchange:")
        for ex, val in sorted(self.summary.get("exchange_totals", {}).items(), key=lambda x: x[1], reverse=True):
            pct = (val / self.summary['total_value_usd'] * 100) if self.summary['total_value_usd'] > 0 else 0
            print(f"   {ex}: ${val:,.2f} ({pct:.1f}%)")


def main():
    parser = argparse.ArgumentParser(description="Export crypto portfolio to various formats")
    parser.add_argument("-f", "--format", choices=["csv", "excel", "json", "all"], 
                        default="csv", help="Export format (default: csv)")
    parser.add_argument("-o", "--output", help="Output filename (auto-generated if not specified)")
    parser.add_argument("--include-staking", action="store_true", help="Include staking details")
    parser.add_argument("--summary", action="store_true", help="Print summary to console")
    
    args = parser.parse_args()
    
    exporter = PortfolioExporter()
    connected = exporter.connect_all()
    
    if connected == 0:
        print("❌ No exchanges connected. Configure API keys in .env file.")
        return
    
    print(f"✓ Connected to {connected} exchange(s)")
    print("📥 Fetching portfolio data...")
    
    exporter.fetch_all_data(include_staking=args.include_staking)
    
    if args.summary:
        exporter.print_summary()
    
    if args.format == "csv" or args.format == "all":
        exporter.export_csv(args.output if args.format == "csv" else None)
    
    if args.format == "excel" or args.format == "all":
        exporter.export_excel(args.output if args.format == "excel" else None)
    
    if args.format == "json" or args.format == "all":
        exporter.export_json(args.output if args.format == "json" else None)


if __name__ == "__main__":
    main()
